# excel_processor_optimized.py
# 大规模数据处理优化版本

import os
import json
import yaml
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Generator
from dataclasses import dataclass
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import gc
import time
from io import BytesIO
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ProcessingConfig:
    split_field: str = ""
    keep_fields: Dict[str, List[str]] = None  # 支持多sheet字段配置
    sort_fields: List[str] = None
    output_dir: str = "output"
    sheet_name: str = "Sheet1"
    selected_sheets: List[str] = None  # 新增：用户选择的要处理的sheet列表
    preserve_format: bool = True
    custom_groups: Dict[str, List[str]] = None
    batch_size: int = 1000  # 批处理大小
    max_workers: int = 4    # 最大线程数
    memory_limit_mb: int = 512  # 内存限制(MB)
    
    def post_init(self):
        if self.keep_fields is None:
            self.keep_fields = {}
        if self.sort_fields is None:
            self.sort_fields = []
        if self.custom_groups is None:
            self.custom_groups = {}
        if self.selected_sheets is None:
            self.selected_sheets = []

class MemoryManager:
    """内存管理器，监控和控制内存使用"""
    
    def __init__(self, limit_mb: int = 512):
        self.limit_bytes = limit_mb * 1024 * 1024
        self._lock = threading.Lock()
        self._psutil_available = False
        
        # 尝试导入psutil
        try:
            import psutil
            self._psutil_available = True
        except ImportError:
            logger.warning("psutil模块不可用，内存监控功能将被禁用")
    
    def check_memory(self) -> bool:
        """检查内存使用是否超限"""
        if not self._psutil_available:
            return True  # 如果psutil不可用，总是返回True（不限制）
        
        try:
            import psutil
            process = psutil.Process()
            memory_info = process.memory_info()
            return memory_info.rss < self.limit_bytes
        except Exception as e:
            logger.warning(f"内存检查失败: {e}")
            return True
    
    def force_gc(self):
        """强制垃圾回收"""
        gc.collect()
    
    def get_memory_usage(self) -> float:
        """获取当前内存使用量(MB)"""
        if not self._psutil_available:
            return 0.0  # 如果psutil不可用，返回0
        
        try:
            import psutil
            process = psutil.Process()
            memory_info = process.memory_info()
            return memory_info.rss / 1024 / 1024
        except Exception as e:
            logger.warning(f"获取内存使用量失败: {e}")
            return 0.0

class ProgressTracker:
    """进度跟踪器"""
    
    def __init__(self, total_steps: int, description: str = "处理中"):
        self.total_steps = total_steps
        self.current_step = 0
        self.description = description
        self.start_time = time.time()
        self._lock = threading.Lock()
    
    def update(self, steps: int = 1):
        """更新进度"""
        with self._lock:
            self.current_step += steps
            elapsed = time.time() - self.start_time
            if self.current_step > 0:
                eta = (elapsed / self.current_step) * (self.total_steps - self.current_step)
                logger.info(f"{self.description}: {self.current_step}/{self.total_steps} "
                           f"({self.current_step/self.total_steps*100:.1f}%) "
                           f"ETA: {eta:.1f}s")
    
    def complete(self):
        """完成处理"""
        elapsed = time.time() - self.start_time
        logger.info(f"{self.description} 完成，耗时: {elapsed:.2f}秒")

class OptimizedExcelProcessor:
    """优化版Excel处理器，支持大规模数据处理"""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.memory_manager = MemoryManager(config.memory_limit_mb)
        self._format_cache = {}  # 格式缓存
        self._workbook_cache = {}  # 工作簿缓存
    
    def read_excel_chunked(self, file_path: str, sheet_name: str = None, 
                          chunk_size: int = None) -> Generator[pd.DataFrame, None, None]:
        """分块读取Excel文件，减少内存占用"""
        chunk_size = chunk_size or self.config.batch_size
        
        # 使用openpyxl的read_only模式
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[sheet_name or self.config.sheet_name]
        
        # 获取表头
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        # 分块读取数据
        current_chunk = []
        for row in sheet.iter_rows(min_row=2):
            if len(current_chunk) >= chunk_size:
                df_chunk = pd.DataFrame(current_chunk, columns=headers)
                yield df_chunk
                current_chunk = []
                self.memory_manager.force_gc()
            
            row_data = [cell.value for cell in row]
            current_chunk.append(row_data)
        
        # 返回最后一块
        if current_chunk:
            df_chunk = pd.DataFrame(current_chunk, columns=headers)
            yield df_chunk
        
        wb.close()
    
    def read_excel_optimized(self, file_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, openpyxl.Workbook]:
        """优化版Excel读取，支持大文件"""
        sheet_name = sheet_name or self.config.sheet_name
        
        # 检查文件大小
        file_size = os.path.getsize(file_path) / 1024 / 1024  # MB
        logger.info(f"文件大小: {file_size:.2f}MB")
        
        if file_size > 50:  # 大文件使用分块读取
            logger.info("检测到大文件，使用分块读取模式")
            return self._read_large_excel(file_path, sheet_name)
        else:
            # 小文件直接读取
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            wb = openpyxl.load_workbook(file_path)
            return df, wb
    
    def _read_large_excel(self, file_path: str, sheet_name: str) -> Tuple[pd.DataFrame, openpyxl.Workbook]:
        """读取大Excel文件"""
        chunks = []
        for chunk in self.read_excel_chunked(file_path, sheet_name):
            chunks.append(chunk)
            if not self.memory_manager.check_memory():
                logger.warning("内存使用接近限制，强制垃圾回收")
                self.memory_manager.force_gc()
        
        df = pd.concat(chunks, ignore_index=True)
        wb = openpyxl.load_workbook(file_path)
        return df, wb
    
    def copy_cell_format_optimized(self, source_cell, target_cell):
        """优化版格式复制，使用缓存"""
        if not self.config.preserve_format:
            return
        
        # 使用缓存键
        cache_key = (id(source_cell.font), id(source_cell.fill), id(source_cell.border), 
                    id(source_cell.alignment), source_cell.number_format)
        
        if cache_key in self._format_cache:
            # 使用缓存的格式对象
            cached_format = self._format_cache[cache_key]
            target_cell.font = cached_format['font']
            target_cell.fill = cached_format['fill']
            target_cell.border = cached_format['border']
            target_cell.alignment = cached_format['alignment']
            target_cell.number_format = cached_format['number_format']
            return
        
        # 创建新的格式对象
        font = copy(source_cell.font) if source_cell.font else None
        fill = copy(source_cell.fill) if source_cell.fill else None
        border = copy(source_cell.border) if source_cell.border else None
        alignment = copy(source_cell.alignment) if source_cell.alignment else None
        number_format = source_cell.number_format
        
        # 缓存格式对象
        self._format_cache[cache_key] = {
            'font': font,
            'fill': fill,
            'border': border,
            'alignment': alignment,
            'number_format': number_format
        }
        
        # 应用格式
        target_cell.font = font
        target_cell.fill = fill
        target_cell.border = border
        target_cell.alignment = alignment
        target_cell.number_format = number_format
        
        # 复制超链接
        if source_cell.hyperlink:
            target_cell.hyperlink = source_cell.hyperlink
    
    def write_excel_with_format_optimized(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                                        output_path: str, sheet_name: str = "Sheet1"):
        """优化版Excel写入，支持大文件"""
        source_ws = wb[sheet_name]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        
        # 复制列宽和行高
        for col_letter, dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width
        for row_idx, dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = dim.height
        
        # 复制合并单元格
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        
        # 获取表头
        header = [cell.value for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
        col_map = {col: idx for idx, col in enumerate(header)}
        
        # 批量写入表头
        for c, v in enumerate(df.columns, 1):
            src_cell = source_ws.cell(row=1, column=col_map[v]+1)
            tgt_cell = new_ws.cell(row=1, column=c, value=v)
            self.copy_cell_format_optimized(src_cell, tgt_cell)
        
        # 批量写入数据行
        batch_size = self.config.batch_size
        total_rows = len(df)
        
        for batch_start in range(0, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            batch_df = df.iloc[batch_start:batch_end]
            
            for r, row in enumerate(batch_df.itertuples(index=False), batch_start + 2):
                for c, v in enumerate(row, 1):
                    # 使用第一行作为格式模板，避免逐行查找
                    src_cell = source_ws.cell(row=2, column=c)
                    tgt_cell = new_ws.cell(row=r, column=c, value=v)
                    self.copy_cell_format_optimized(src_cell, tgt_cell)
            
            # 定期清理内存
            if batch_start % (batch_size * 10) == 0:
                self.memory_manager.force_gc()
        
        # 使用内存流保存，避免临时文件
        with BytesIO() as buffer:
            new_wb.save(buffer)
            buffer.seek(0)
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())
        
        new_wb.close()
    
    def split_excel_optimized(self, input_file: str, sheet_name: str = None, 
                            progress_callback=None) -> List[str]:
        """优化版Excel拆分，支持大文件和多sheet"""
        logger.info(f"开始处理文件: {input_file}")
        
        # 读取sheet名时可用read_only=True，但后续格式复制必须用默认模式
        wb = openpyxl.load_workbook(input_file)  # 不加read_only=True，保证格式属性可用
        
        # 确定要处理的sheet列表
        sheets_to_process = []
        if self.config.selected_sheets:
            # 使用用户选择的sheet列表
            sheets_to_process = [sheet for sheet in self.config.selected_sheets if sheet in wb.sheetnames]
            if not sheets_to_process:
                raise ValueError(f"用户选择的sheet都不存在于文件中: {self.config.selected_sheets}")
        else:
            # 兼容旧版本，使用单个sheet
            use_sheet = sheet_name or self.config.sheet_name or wb.sheetnames[0]
            if use_sheet not in wb.sheetnames:
                use_sheet = wb.sheetnames[0]
            sheets_to_process = [use_sheet]
        
        logger.info(f"将处理以下sheet: {sheets_to_process}")
        
        all_output_files = []
        
        # 对每个选中的sheet进行处理
        for current_sheet in sheets_to_process:
            try:
                logger.info(f"正在处理sheet: {current_sheet}")
                
                # 只读取当前要处理的sheet
                df = pd.read_excel(input_file, sheet_name=current_sheet, header=0)
                
                # 检查拆分字段是否存在
                if self.config.split_field not in df.columns:
                    logger.warning(f"拆分字段 '{self.config.split_field}' 在sheet '{current_sheet}' 中不存在，跳过该sheet")
                    continue
                
                # 应用字段筛选
                if self.config.keep_fields and current_sheet in self.config.keep_fields:
                    available_fields = [col for col in self.config.keep_fields[current_sheet] if col in df.columns]
                    df = df[available_fields]
                
                # 应用排序
                if self.config.sort_fields:
                    sort_fields = [col for col in self.config.sort_fields if col in df.columns]
                    if sort_fields:
                        df = df.sort_values(by=sort_fields)
                
                # 检查自定义分组
                if self.config.custom_groups:
                    sheet_output_files = self.split_excel_with_groups_optimized(df, wb, current_sheet, progress_callback)
                else:
                    sheet_output_files = self.split_excel_traditional_optimized(df, wb, current_sheet, progress_callback)
                
                all_output_files.extend(sheet_output_files)
                
            except Exception as e:
                logger.error(f"处理sheet '{current_sheet}' 时出错: {e}")
                continue
        
        return all_output_files
    
    def split_excel_traditional_optimized(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                                        sheet_name: str, progress_callback=None) -> List[str]:
        """优化版传统拆分模式"""
        split_values = df[self.config.split_field].unique()
        output_files = []
        
        progress = ProgressTracker(len(split_values), "拆分处理")
        
        # 使用线程池并行处理
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            futures = []
            
            for value in split_values:
                future = executor.submit(
                    self._process_single_split, df, wb, sheet_name, value
                )
                futures.append(future)
            
            # 收集结果
            for future in as_completed(futures):
                try:
                    output_file = future.result()
                    if output_file:
                        output_files.append(output_file)
                    progress.update()
                    if progress_callback:
                        progress_callback(progress.current_step, progress.total_steps)
                except Exception as e:
                    logger.error(f"处理拆分值时出错: {e}")
        
        progress.complete()
        return output_files
    
    def _process_single_split(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                            sheet_name: str, value) -> str:
        """处理单个拆分值"""
        subset = df[df[self.config.split_field] == value]
        if subset.empty:
            return None
        
        safe_value = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
        output_file = self.output_dir / f"{self.config.split_field}-{safe_value}.xlsx"
        
        self.write_excel_with_format_optimized(subset, wb, str(output_file), sheet_name)
        return str(output_file)
    
    def split_excel_with_groups_optimized(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                                        sheet_name: str, progress_callback=None) -> List[str]:
        """优化版自定义分组拆分"""
        output_files = []
        
        # 验证分组配置
        all_group_values = set()
        for group_values in self.config.custom_groups.values():
            all_group_values.update(group_values)
        
        split_values = set(df[self.config.split_field].astype(str).unique())
        unassigned = split_values - all_group_values
        
        if unassigned:
            logger.warning(f"以下字段值未分配到任何分组: {unassigned}")
        
        progress = ProgressTracker(len(self.config.custom_groups), "分组处理")
        
        # 并行处理分组
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            futures = []
            
            for group_name, group_values in self.config.custom_groups.items():
                if not group_values:
                    continue
                
                future = executor.submit(
                    self._process_single_group, df, wb, sheet_name, group_name, group_values
                )
                futures.append(future)
            
            # 收集结果
            for future in as_completed(futures):
                try:
                    output_file = future.result()
                    if output_file:
                        output_files.append(output_file)
                    progress.update()
                    if progress_callback:
                        progress_callback(progress.current_step, progress.total_steps)
                except Exception as e:
                    logger.error(f"处理分组时出错: {e}")
        
        progress.complete()
        return output_files
    
    def _process_single_group(self, df: pd.DataFrame, wb: openpyxl.Workbook, 
                            sheet_name: str, group_name: str, group_values: List[str]) -> str:
        """处理单个分组"""
        subset = df[df[self.config.split_field].astype(str).isin(group_values)]
        
        if subset.empty:
            logger.warning(f"分组 '{group_name}' 没有匹配的数据")
            return None
        
        safe_group_name = group_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        output_file = self.output_dir / f"{safe_group_name}.xlsx"
        
        self.write_excel_with_format_optimized(subset, wb, str(output_file), sheet_name)
        logger.info(f"分组 '{group_name}' 完成，包含 {len(subset)} 行数据")
        return str(output_file)
    
    def merge_excel_files_optimized(self, input_files: list, output_file: str, 
                                  progress_callback=None) -> str:
        """优化版Excel合并"""
        logger.info(f"开始合并 {len(input_files)} 个文件")
        
        progress = ProgressTracker(len(input_files), "文件合并")
        
        all_data = []
        reference_wb = None
        
        for i, file_path in enumerate(input_files):
            try:
                # 合并时也用默认模式，保证格式属性可用
                wb = openpyxl.load_workbook(file_path)
                first_sheet = wb.sheetnames[0]
                df = pd.read_excel(file_path, sheet_name=first_sheet)
                if reference_wb is None:
                    reference_wb = wb
                
                # 应用字段筛选
                if self.config.keep_fields:
                    if first_sheet in self.config.keep_fields:
                        available_fields = [col for col in self.config.keep_fields[first_sheet] if col in df.columns]
                        df = df[available_fields]
                
                all_data.append(df)
                progress.update()
                if progress_callback:
                    progress_callback(progress.current_step, progress.total_steps)
                
            except Exception as e:
                logger.error(f"读取文件 {file_path} 时出错: {e}")
                continue
        
        if not all_data:
            raise ValueError("没有成功读取任何文件")
        
        # 合并数据
        merged_df = pd.concat(all_data, ignore_index=True)
        
        # 应用排序
        if self.config.sort_fields:
            sort_fields = [col for col in self.config.sort_fields if col in merged_df.columns]
            if sort_fields:
                merged_df = merged_df.sort_values(by=sort_fields)
        
        # 写入结果
        sheet_name = self.config.sheet_name
        if sheet_name not in reference_wb.sheetnames:
            sheet_name = reference_wb.sheetnames[0]
        
        output_path = self.output_dir / output_file
        self.write_excel_with_format_optimized(merged_df, reference_wb, str(output_path), sheet_name)
        
        progress.complete()
        return str(output_path)
    
    def create_zip_archive(self, file_paths: List[str], zip_name: str) -> str:
        """创建ZIP压缩包"""
        zip_path = self.output_dir / zip_name
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in file_paths:
                if os.path.exists(file_path):
                    zipf.write(file_path, arcname=os.path.basename(file_path))
        
        return str(zip_path)
    
    def cleanup_cache(self):
        """清理缓存"""
        self._format_cache.clear()
        self._workbook_cache.clear()
        self.memory_manager.force_gc()

def load_config_optimized(config_file: str) -> ProcessingConfig:
    """加载优化版配置"""
    config_path = Path(config_file)
    if not config_path.exists():
        raise FileNotFoundError(f"配置文件不存在: {config_file}")
    
    with open(config_path, 'r', encoding='utf-8') as f:
        if config_path.suffix.lower() == '.json':
            config_data = json.load(f)
        elif config_path.suffix.lower() in ['.yml', '.yaml']:
            config_data = yaml.safe_load(f)
        else:
            raise ValueError(f"不支持的配置文件格式: {config_path.suffix}")
    
    return ProcessingConfig(**config_data) 